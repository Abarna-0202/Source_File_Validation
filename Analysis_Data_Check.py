import pandas as pd
from openpyxl import load_workbook
import os

def column_letter_to_number(letter):
    """Convert Excel column letter (e.g., 'A') to 1-based number."""
    letter = letter.upper()
    return sum((ord(char) - 64) * 26**i for i, char in enumerate(reversed(letter)))

def map_columns(sample_file, mapping_file, output_file):
    try:
        sample_data = pd.read_excel(sample_file, sheet_name='Empty Columns Summary')
    except ValueError as e:
        print(f"Error: {e}")
        workbook = load_workbook(sample_file)
        sheet_names = workbook.sheetnames
        print(f"Available sheet names in the file: {sheet_names}")
        return
    mapping_data = pd.read_excel(mapping_file, sheet_name=None, header=1)
    sample_data['Mapped Value'] = None
    for index, row in sample_data.iterrows():
        entity = row['Customer entity name'] 
        column_letter = row['Column Name']
        column_number = column_letter_to_number(column_letter)
        
        if entity in mapping_data:
            mapping_sheet = mapping_data[entity]
            mapping_sheet.columns = mapping_sheet.columns.astype(str).str.strip()
            position_col = None
            for col in mapping_sheet.columns:
                if 'Position' in col:
                    position_col = col
                    break
            
            if position_col:
                mapping_sheet[position_col] = pd.to_numeric(mapping_sheet[position_col], errors='coerce')
                mapped_row = mapping_sheet[mapping_sheet[position_col] == column_number]
                
                if not mapped_row.empty:
                    if 'Column' in mapped_row.columns:
                        mapped_value = mapped_row['Column'].values[0]
                    elif 'RM Column' in mapped_row.columns:
                        mapped_value = mapped_row['RM Column'].values[0]
                    else:
                        mapped_value = None
                    sample_data.at[index, 'Mapped Value'] = mapped_value
            else:
                print(f"'Position' column missing in sheet '{entity}'. Columns: {mapping_sheet.columns}")
    sample_data.to_excel(output_file, index=False)
    print(f"Updated file saved: {output_file}")
    if os.path.exists(output_file):
        print(f"File exists: {output_file}")
    else:
        print(f"File does not exist: {output_file}")

def combine_and_group_sheets(updated_sample_file, output_combined_file):
    if not os.path.exists(updated_sample_file):
        print(f"Error: File does not exist: {updated_sample_file}")
        return
    updated_sample_data = pd.read_excel(updated_sample_file, sheet_name=None)
    combined_rows = []
    for sheet_name, sheet_data in updated_sample_data.items():
        grouped = sheet_data.groupby(
            ['Customer entity name', 'Column Name', 'Mapped Value'], as_index=False
        ).agg({
            'Status': 'first',
            'Empty Count': 'sum'
        })
        combined_rows.append(grouped)
    combined_df = pd.concat(combined_rows, ignore_index=True)
    combined_df = combined_df.sort_values(by='Status', key=lambda x: x == 'Pass')
    combined_df.to_excel(output_combined_file, index=False)
    print(f"Combined, grouped, and sorted data saved to {output_combined_file}")

def final_output(file_with_3_sheets, file_with_1_sheet):
    if not os.path.exists(file_with_1_sheet):
        print(f"Error: File does not exist: {file_with_1_sheet}")
        return
    book = load_workbook(file_with_3_sheets)
    if len(book.worksheets) < 3:
        print(f"Error: The workbook '{file_with_3_sheets}' does not have at least 3 sheets.")
        print(f"Available sheets: {book.sheetnames}")
        return
    new_data = pd.read_excel(file_with_1_sheet)
    third_sheet = book.worksheets[2]
    for row in third_sheet.iter_rows(min_row=2, max_row=third_sheet.max_row, max_col=third_sheet.max_column):
        for cell in row:
            cell.value = None
    for index, row in new_data.iterrows():
        for col_num, value in enumerate(row, 1):
            third_sheet.cell(row=index+2, column=col_num, value=value)
    book.save(file_with_3_sheets)
    print(f"Data in third sheet replaced and saved to {file_with_3_sheets}")
    
def process_and_summarize_data(mapping_file, output_file_master, output_file_mapped, output_combined_file):
    map_columns(output_file_master, mapping_file, output_file_mapped)
    combine_and_group_sheets(output_file_mapped, output_combined_file)
    combined_data = pd.read_excel(output_combined_file)
    combined_data.to_excel(output_combined_file, index=False)
    print(f"Final consolidated report saved to {output_combined_file}")